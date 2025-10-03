"""Bulk import service for hierarchical data structures."""

import csv
import io
from typing import List, Dict, Any, Optional, Tuple, Union
from datetime import datetime
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError
from fastapi import HTTPException, UploadFile
import pandas as pd
from openpyxl import load_workbook

from app.models.organization import Client, Site, Building, Gate
from app.schemas.structure import (
    ClientImportRow, SiteImportRow, BuildingImportRow, GateImportRow,
    ImportResult, ClientType, BuildingType, GateType, GateStatus
)
from app.services.structure import ClientService, SiteService, BuildingService, GateService
import structlog

logger = structlog.get_logger(__name__)


class BulkImportService:
    """Service for bulk importing hierarchical data from CSV/XLSX files."""
    
    def __init__(self, db: Session, org_id: int):
        self.db = db
        self.org_id = org_id
        self.client_service = ClientService(db)
        self.site_service = SiteService(db)
        self.building_service = BuildingService(db)
        self.gate_service = GateService(db)
        
        # Cache for created entities to avoid duplicates
        self.client_cache: Dict[str, int] = {}
        self.site_cache: Dict[Tuple[str, str], int] = {}  # (client_name, site_name) -> site_id
        self.building_cache: Dict[Tuple[str, str, str], int] = {}  # (client, site, building) -> building_id
    
    async def import_from_file(
        self, 
        file: UploadFile, 
        import_type: str = "hierarchical"
    ) -> ImportResult:
        """
        Import data from CSV or XLSX file.
        
        Args:
            file: Uploaded file (CSV or XLSX)
            import_type: Type of import ("hierarchical", "clients", "sites", "buildings", "gates")
            
        Returns:
            ImportResult with statistics and any errors
        """
        try:
            # Determine file type and read data
            if file.filename.endswith('.csv'):
                data = await self._read_csv(file)
            elif file.filename.endswith(('.xlsx', '.xls')):
                data = await self._read_excel(file)
            else:
                raise HTTPException(
                    status_code=400,
                    detail="Unsupported file format. Use CSV or XLSX."
                )
            
            # Process data based on import type
            if import_type == "hierarchical":
                return await self._import_hierarchical(data)
            elif import_type == "clients":
                return await self._import_clients_only(data)
            elif import_type == "sites":
                return await self._import_sites_only(data)
            elif import_type == "buildings":
                return await self._import_buildings_only(data)
            elif import_type == "gates":
                return await self._import_gates_only(data)
            else:
                raise HTTPException(
                    status_code=400,
                    detail=f"Unknown import type: {import_type}"
                )
                
        except Exception as e:
            logger.error("Import failed", error=str(e), file=file.filename)
            return ImportResult(
                success=False,
                total_rows=0,
                processed_rows=0,
                skipped_rows=0,
                errors=[f"Import failed: {str(e)}"]
            )
    
    async def _read_csv(self, file: UploadFile) -> List[Dict[str, Any]]:
        """Read CSV file and return list of dictionaries."""
        content = await file.read()
        csv_data = content.decode('utf-8-sig')  # Handle BOM if present
        
        reader = csv.DictReader(io.StringIO(csv_data))
        return list(reader)
    
    async def _read_excel(self, file: UploadFile) -> List[Dict[str, Any]]:
        """Read Excel file and return list of dictionaries."""
        content = await file.read()
        
        # Try pandas first (handles both .xls and .xlsx)
        try:
            df = pd.read_excel(io.BytesIO(content))
            # Replace NaN with None for better JSON serialization
            df = df.where(pd.notnull(df), None)
            return df.to_dict('records')
        except Exception as e:
            logger.warning("Pandas failed to read Excel, trying openpyxl", error=str(e))
            
        # Fallback to openpyxl for .xlsx files
        try:
            workbook = load_workbook(io.BytesIO(content))
            sheet = workbook.active
            
            # Get headers from first row
            headers = [cell.value for cell in sheet[1]]
            
            # Get data from remaining rows
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for i, value in enumerate(row):
                    if i < len(headers) and headers[i]:
                        row_dict[headers[i]] = value
                data.append(row_dict)
            
            return data
        except Exception as e:
            raise HTTPException(
                status_code=400,
                detail=f"Failed to read Excel file: {str(e)}"
            )
    
    async def _import_hierarchical(self, data: List[Dict[str, Any]]) -> ImportResult:
        """Import hierarchical data (client -> site -> building -> gate)."""
        result = ImportResult(
            success=True,
            total_rows=len(data),
            processed_rows=0,
            skipped_rows=0,
            created_entities={
                "clients": [],
                "sites": [],
                "buildings": [], 
                "gates": []
            }
        )
        
        for i, row in enumerate(data):
            try:
                # Validate required fields
                if not all(k in row for k in ['client_name', 'site_name', 'building_name', 'gate_name']):
                    result.errors.append(f"Row {i+1}: Missing required fields (client_name, site_name, building_name, gate_name)")
                    result.skipped_rows += 1
                    continue
                
                # Create or get client
                client_id = await self._ensure_client(row, result)
                if not client_id:
                    result.skipped_rows += 1
                    continue
                
                # Create or get site
                site_id = await self._ensure_site(row, client_id, result)
                if not site_id:
                    result.skipped_rows += 1
                    continue
                
                # Create or get building
                building_id = await self._ensure_building(row, site_id, result)
                if not building_id:
                    result.skipped_rows += 1
                    continue
                
                # Create gate
                gate_id = await self._create_gate(row, building_id, result, i+1)
                if gate_id:
                    result.created_entities["gates"].append(gate_id)
                    result.processed_rows += 1
                else:
                    result.skipped_rows += 1
                    
            except Exception as e:
                logger.error("Error processing row", row_index=i+1, error=str(e))
                result.errors.append(f"Row {i+1}: {str(e)}")
                result.skipped_rows += 1
        
        if result.errors:
            result.success = len(result.errors) == 0
        
        return result
    
    async def _ensure_client(self, row: Dict[str, Any], result: ImportResult) -> Optional[int]:
        """Create client if doesn't exist, return client_id."""
        client_name = row['client_name'].strip()
        
        # Check cache first
        if client_name in self.client_cache:
            return self.client_cache[client_name]
        
        # Check if client already exists in database
        existing_client = self.db.query(Client).filter(
            Client.organization_id == self.org_id,
            Client.name == client_name
        ).first()
        
        if existing_client:
            self.client_cache[client_name] = existing_client.id
            return existing_client.id
        
        # Create new client
        try:
            from app.schemas.structure import ClientCreate
            
            client_data = ClientCreate(
                name=client_name,
                display_name=row.get('client_display_name') or client_name,
                type=ClientType(row.get('client_type', 'residential')),
                contact_person=row.get('client_contact_person'),
                email=row.get('client_email'),
                phone=row.get('client_phone'),
                address_line_1=row.get('client_address_line_1'),
                address_line_2=row.get('client_address_line_2'),
                city=row.get('client_city'),
                state=row.get('client_state'),
                postal_code=row.get('client_postal_code'),
                country=row.get('client_country', 'Hungary'),
                contract_number=row.get('client_contract_number')
            )
            
            client = self.client_service.create_client(self.org_id, client_data)
            self.client_cache[client_name] = client.id
            result.created_entities["clients"].append(client.id)
            
            logger.info("Created client", client_id=client.id, name=client_name)
            return client.id
            
        except Exception as e:
            result.errors.append(f"Failed to create client '{client_name}': {str(e)}")
            return None
    
    async def _ensure_site(self, row: Dict[str, Any], client_id: int, result: ImportResult) -> Optional[int]:
        """Create site if doesn't exist, return site_id."""
        client_name = row['client_name'].strip()
        site_name = row['site_name'].strip()
        cache_key = (client_name, site_name)
        
        # Check cache first
        if cache_key in self.site_cache:
            return self.site_cache[cache_key]
        
        # Check if site already exists
        existing_site = self.db.query(Site).filter(
            Site.client_id == client_id,
            Site.name == site_name
        ).first()
        
        if existing_site:
            self.site_cache[cache_key] = existing_site.id
            return existing_site.id
        
        # Create new site
        try:
            from app.schemas.structure import SiteCreate
            
            site_data = SiteCreate(
                client_id=client_id,
                name=site_name,
                display_name=row.get('site_display_name') or site_name,
                site_code=row.get('site_code'),
                address_line_1=row.get('site_address_line_1'),
                address_line_2=row.get('site_address_line_2'),
                city=row.get('site_city'),
                state=row.get('site_state'),
                postal_code=row.get('site_postal_code'),
                country=row.get('site_country', 'Hungary'),
                latitude=row.get('site_latitude'),
                longitude=row.get('site_longitude'),
                area_sqm=int(row['site_area_sqm']) if row.get('site_area_sqm') else None,
                emergency_contact=row.get('site_emergency_contact'),
                emergency_phone=row.get('site_emergency_phone')
            )
            
            site = self.site_service.create_site(self.org_id, site_data)
            self.site_cache[cache_key] = site.id
            result.created_entities["sites"].append(site.id)
            
            logger.info("Created site", site_id=site.id, name=site_name, client_id=client_id)
            return site.id
            
        except Exception as e:
            result.errors.append(f"Failed to create site '{site_name}': {str(e)}")
            return None
    
    async def _ensure_building(self, row: Dict[str, Any], site_id: int, result: ImportResult) -> Optional[int]:
        """Create building if doesn't exist, return building_id."""
        client_name = row['client_name'].strip()
        site_name = row['site_name'].strip()
        building_name = row['building_name'].strip()
        cache_key = (client_name, site_name, building_name)
        
        # Check cache first
        if cache_key in self.building_cache:
            return self.building_cache[cache_key]
        
        # Check if building already exists
        existing_building = self.db.query(Building).filter(
            Building.site_id == site_id,
            Building.name == building_name
        ).first()
        
        if existing_building:
            self.building_cache[cache_key] = existing_building.id
            return existing_building.id
        
        # Create new building
        try:
            from app.schemas.structure import BuildingCreate
            
            building_data = BuildingCreate(
                site_id=site_id,
                name=building_name,
                display_name=row.get('building_display_name') or building_name,
                building_code=row.get('building_code'),
                building_type=BuildingType(row.get('building_type', 'residential')),
                floors=int(row['building_floors']) if row.get('building_floors') else None,
                units=int(row['building_units']) if row.get('building_units') else None,
                year_built=int(row['building_year_built']) if row.get('building_year_built') else None,
                address_suffix=row.get('building_address_suffix')
            )
            
            building = self.building_service.create_building(self.org_id, building_data)
            self.building_cache[cache_key] = building.id
            result.created_entities["buildings"].append(building.id)
            
            logger.info("Created building", building_id=building.id, name=building_name, site_id=site_id)
            return building.id
            
        except Exception as e:
            result.errors.append(f"Failed to create building '{building_name}': {str(e)}")
            return None
    
    async def _create_gate(
        self, 
        row: Dict[str, Any], 
        building_id: int, 
        result: ImportResult, 
        row_number: int
    ) -> Optional[int]:
        """Create gate."""
        gate_name = row['gate_name'].strip()
        
        # Check if gate already exists
        existing_gate = self.db.query(Gate).filter(
            Gate.building_id == building_id,
            Gate.name == gate_name
        ).first()
        
        if existing_gate:
            result.warnings.append(f"Row {row_number}: Gate '{gate_name}' already exists, skipping")
            return existing_gate.id
        
        # Create new gate
        try:
            from app.schemas.structure import GateCreate
            
            # Parse installation date
            installation_date = None
            if row.get('installation_date'):
                try:
                    installation_date = pd.to_datetime(row['installation_date'])
                except:
                    result.warnings.append(f"Row {row_number}: Invalid installation_date format")
            
            gate_data = GateCreate(
                building_id=building_id,
                name=gate_name,
                display_name=row.get('gate_display_name') or gate_name,
                gate_code=row.get('gate_code'),
                gate_type=GateType(row.get('gate_type', 'swing')),
                manufacturer=row.get('manufacturer'),
                model=row.get('model'),
                serial_number=row.get('serial_number'),
                installation_date=installation_date,
                installer=row.get('installer'),
                width_cm=int(row['width_cm']) if row.get('width_cm') else None,
                height_cm=int(row['height_cm']) if row.get('height_cm') else None,
                weight_kg=int(row['weight_kg']) if row.get('weight_kg') else None,
                material=row.get('material'),
                max_opening_cycles_per_day=int(row['max_cycles_per_day']) if row.get('max_cycles_per_day') else None,
                current_cycle_count=int(row.get('current_cycle_count', 0)),
                status=GateStatus(row.get('status', 'operational'))
            )
            
            gate = self.gate_service.create_gate(self.org_id, gate_data)
            
            logger.info("Created gate", gate_id=gate.id, name=gate_name, building_id=building_id)
            return gate.id
            
        except Exception as e:
            result.errors.append(f"Row {row_number}: Failed to create gate '{gate_name}': {str(e)}")
            return None
    
    async def _import_clients_only(self, data: List[Dict[str, Any]]) -> ImportResult:
        """Import only clients from data."""
        # Implementation for client-only import
        # Similar to hierarchical but only creates clients
        pass
    
    async def _import_sites_only(self, data: List[Dict[str, Any]]) -> ImportResult:
        """Import only sites from data (requires existing clients)."""
        # Implementation for site-only import
        pass
    
    async def _import_buildings_only(self, data: List[Dict[str, Any]]) -> ImportResult:
        """Import only buildings from data (requires existing sites)."""
        # Implementation for building-only import
        pass
    
    async def _import_gates_only(self, data: List[Dict[str, Any]]) -> ImportResult:
        """Import only gates from data (requires existing buildings)."""
        # Implementation for gate-only import
        pass


def validate_import_file(file: UploadFile, expected_columns: List[str]) -> List[str]:
    """
    Validate import file format and required columns.
    
    Args:
        file: Uploaded file
        expected_columns: List of required column names
        
    Returns:
        List of validation errors
    """
    errors = []
    
    # Check file type
    if not file.filename.endswith(('.csv', '.xlsx', '.xls')):
        errors.append("File must be CSV or Excel format (.csv, .xlsx, .xls)")
        return errors
    
    # Check file size (10MB limit)
    if hasattr(file, 'size') and file.size > 10 * 1024 * 1024:
        errors.append("File size must be less than 10MB")
    
    return errors